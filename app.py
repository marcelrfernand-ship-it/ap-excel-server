import os
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import base64, io, json, traceback
import requests as req_asana
from template_b64 import TEMPLATE_B64

app = Flask(__name__)
CORS(app)

ASANA_TOKEN = os.environ.get('ASANA_TOKEN', '')
ASANA_WORKSPACE = '1202781323743076'
ASANA_BASE = 'https://app.asana.com/api/1.0'
ASANA_HEADERS = {'Authorization': f'Bearer {ASANA_TOKEN}', 'Content-Type': 'application/json'}
EMAIL_GID_CACHE = {}

ANTHROPIC_KEY = os.environ.get('ANTHROPIC_KEY', '')

def sc(ws, addr, val):
    if val is None or val == '': return
    if addr not in ws: ws[addr] = None
    cell = ws[addr]
    if isinstance(val, (int, float)):
        cell.value = val
    else:
        cell.value = str(val)

def get_user_gid(email):
    if email in EMAIL_GID_CACHE: return EMAIL_GID_CACHE[email]
    try:
        r = req_asana.get(f'{ASANA_BASE}/workspaces/{ASANA_WORKSPACE}/typeahead',
            params={'resource_type':'user','query':email}, headers=ASANA_HEADERS)
        data = r.json().get('data',[])
        if data:
            EMAIL_GID_CACHE[email] = data[0]['gid']
            return data[0]['gid']
    except:
        pass
    return None

@app.route('/', methods=['GET'])
def index():
    return jsonify({'status': 'ok', 'service': 'AP Excel Generator'})

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'AP-WDC Server'})

# ── CLAUDE PROXY ──
@app.route('/claude', methods=['POST', 'OPTIONS'])
def claude_proxy():
    if request.method == 'OPTIONS':
        res = jsonify({'ok': True})
        res.headers.add('Access-Control-Allow-Origin', '*')
        res.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        res.headers.add('Access-Control-Allow-Methods', 'POST')
        return res
    try:
        if not ANTHROPIC_KEY:
            res = jsonify({'error': 'ANTHROPIC_KEY nao configurada no servidor'})
            res.headers.add('Access-Control-Allow-Origin', '*')
            return res, 500

        d = request.get_json(force=True)
        r = req_asana.post(
            'https://api.anthropic.com/v1/messages',
            json={
                'model': d.get('model', 'claude-sonnet-4-20250514'),
                'max_tokens': d.get('max_tokens', 1000),
                'messages': d.get('messages', [])
            },
            headers={
                'x-api-key': ANTHROPIC_KEY,
                'anthropic-version': '2023-06-01',
                'Content-Type': 'application/json'
            },
            timeout=60
        )
        res = jsonify(r.json())
        res.headers.add('Access-Control-Allow-Origin', '*')
        return res, r.status_code
    except Exception as e:
        res = jsonify({'error': str(e)})
        res.headers.add('Access-Control-Allow-Origin', '*')
        return res, 500

@app.route('/gerar-excel', methods=['POST'])
def gerar_excel():
    try:
        d = request.get_json(force=True)
        h = d.get('header', {})
        p = d.get('perfil', {})

        template_bytes = base64.b64decode(TEMPLATE_B64)
        wb = load_workbook(io.BytesIO(template_bytes))

        ws = wb['🏠 Perfil']
        sc(ws, 'B5', h.get('nome'))
        sc(ws, 'G5', h.get('kam'))
        sc(ws, 'J5', h.get('inside'))
        sc(ws, 'B8', h.get('data'))
        sc(ws, 'E8', h.get('status'))
        sc(ws, 'B13', p.get('receita'))
        sc(ws, 'F13', p.get('assin'))
        sc(ws, 'I13', p.get('cidades'))
        sc(ws, 'L13', p.get('taxa'))
        sc(ws, 'B16', p.get('fat'))
        sc(ws, 'F16', p.get('pos'))
        sc(ws, 'B19', p.get('perspectiva'))
        if p.get('rwdc'): sc(ws, 'B36', float(p['rwdc']))
        if p.get('rpot'): sc(ws, 'C36', float(p['rpot']))
        sc(ws, 'B41', p.get('sativas'))
        sc(ws, 'B46', p.get('salvo'))
        sc(ws, 'B51', p.get('compra'))
        sc(ws, 'B56', p.get('conc'))
        sc(ws, 'B61', p.get('obs'))

        ws = wb['👥 Contatos']
        for i, ct in enumerate(d.get('contatos', [])[:20]):
            r = 7 + i
            sc(ws, f'B{r}', ct.get('nome'))
            sc(ws, f'C{r}', ct.get('cargo'))
            sc(ws, f'D{r}', ct.get('exec'))
            sc(ws, f'E{r}', ct.get('c'))
            sc(ws, f'F{r}', ct.get('a'))
            sc(ws, f'G{r}', ct.get('inf'))
            sc(ws, f'H{r}', ct.get('tel'))
            sc(ws, f'I{r}', ct.get('email'))
            sc(ws, f'J{r}', ct.get('obs'))

        ws = wb['📁 Projetos']
        for i, pr in enumerate(d.get('proje
